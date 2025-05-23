# import streamlit as st
# import joblib
# import numpy as np


# # Load models
# chest_model = joblib.load("AB_Gait_Chest.joblib")
# hand_model = joblib.load("AB_Gait_Hand.joblib")

# # Function to preprocess input
# def parse_input(input_str):
#     try:
#         values = list(map(float, input_str.strip().split(',')))
#         if len(values) != 300:
#             return None, "Input must contain exactly 300 comma-separated values."
#         return np.array(values).reshape(1, -1), None
#     except ValueError:
#         return None, "Please enter only numeric values separated by commas."

# # UI
# st.title("Gait Analysis for Parkinson Detection")
# st.markdown("This app analyzes gait data from **Left Hand** and **Right Hand** sensors to detect Parkinson’s Disease.")

# # --- Chest Section ---
# st.header("📍 Left Hand Gait Data")
# chest_input = st.text_area("Enter 300 comma-separated values for chest sensor:")

# if st.button("Analyze Left Hand"):
#     chest_array, error = parse_input(chest_input)
#     if error:
#         st.error(error)
#     else:
#         prediction = chest_model.predict(chest_array)[0]
#         result = "🟢 Healthy" if prediction == 0 else "🔴 Parkinson"
#         st.success(f"**Left Hand Gait Prediction:** {result}")

#                 # Save to session
#         st.session_state["gait_chest_result"] = result
#         st.session_state["gait_chest_data"] = chest_array

# # --- Divider ---
# st.markdown("---")

# # --- Hand Section ---
# st.header("📍 Right Hand Gait Data")
# hand_input = st.text_area("Enter 300 comma-separated values for hand sensor:")

# if st.button("Analyze Right Hand"):
#     hand_array, error = parse_input(hand_input)
#     if error:
#         st.error(error)
#     else:
#         prediction = hand_model.predict(hand_array)[0]
#         result = "🟢 Healthy" if prediction == 0 else "🔴 Parkinson"
#         st.success(f"**Right Hand Gait Prediction:** {result}")

#          # Save to session
#         st.session_state["gait_hand_result"] = result
#         st.session_state["gait_hand_data"] = hand_array


import streamlit as st
import joblib
import numpy as np
from PIL import Image
import io
from openpyxl import load_workbook

# Load models
chest_model = joblib.load("AB_Gait_Chest.joblib")
hand_model = joblib.load("AB_Gait_Hand.joblib")


# Extract a specific cell using openpyxl
def extract_cell_value(file, cell_address):
    try:
        wb = load_workbook(filename=file, read_only=True)
        ws = wb.active
        value = ws[cell_address].value
        return str(value)
    except Exception as e:
        return None

# Function to preprocess input
def parse_input(input_str):
    try:
        values = list(map(float, input_str.strip().split(',')))
        if len(values) != 300:
            return None, "Input must contain exactly 300 comma-separated values."
        return np.array(values).reshape(1, -1), None
    except ValueError:
        return None, "Please enter only numeric values separated by commas."
    

# Function to convert image to bytes
def image_to_bytes(image):
    byte_arr = io.BytesIO()
    image.save(byte_arr, format="PNG")  # You can change the format if needed (e.g., "JPEG")
    byte_arr.seek(0)
    return byte_arr.read()

# Initialize session state variables if they don't exist
if "gait_chest_result" not in st.session_state:
    st.session_state["gait_chest_result"] = None
if "gait_hand_result" not in st.session_state:
    st.session_state["gait_hand_result"] = None
if "gait_chest_data" not in st.session_state:
    st.session_state["gait_chest_data"] = None
if "gait_hand_data" not in st.session_state:
    st.session_state["gait_hand_data"] = None
if "left_hand_image" not in st.session_state:
    st.session_state["left_hand_image"] = None
if "right_hand_image" not in st.session_state:
    st.session_state["right_hand_image"] = None

# UI
st.title("Gait Analysis for Parkinson Detection")
st.markdown("This app analyzes gait data from **Left Hand** and **Right Hand** sensors to detect Parkinson's Disease.")

# --- Left Hand Section ---
st.header("📍 Left Hand Gait Data") 

## refer chest as left hand (VV imp)
left_excel = st.file_uploader("Upload Left Hand Excel File (.xlsx)", type=["xlsx"], key="left_xlsx")
left_cell = "I2"

chest_input = ""
if left_excel and left_cell:
    extracted = extract_cell_value(left_excel, left_cell)
    if extracted:
        chest_input = extracted
        st.success("✅ Data extracted from cell.")
    else:
        st.error("❌ Could not extract from specified cell.")

chest_input = st.text_area("Comma-separated 300 values (Left Hand) for left hand sensor", value=chest_input, height=150)

# chest_input = st.text_area("Enter 300 comma-separated values for left hand sensor:")


# Left Hand Image Upload
st.subheader("Upload Left Hand Plot")
left_hand_img = st.file_uploader("Choose an image of left hand gait analysis", type=["jpg", "jpeg", "png"], key="left_hand")
if left_hand_img is not None:
    image = Image.open(left_hand_img)
    st.image(image, caption="Uploaded Left Hand Plot", width=300)

    # Convert image to bytes and store in session state
    left_hand_image_bytes = image_to_bytes(image)
    st.session_state["left_hand_image_bytes"] = left_hand_image_bytes

if st.button("Analyze Left Hand"):
    chest_array, error = parse_input(chest_input)
    if error:
        st.error(error)
    else:
        prediction = chest_model.predict(chest_array)[0]
        result = "🟢 Healthy" if prediction == 0 else "🔴 Parkinson"
        st.success(f"**Left Hand Gait Prediction:** {result}")

        # Save to session
        st.session_state["gait_chest_result"] = result
        st.session_state["gait_chest_data"] = chest_array

# --- Divider ---
st.markdown("---")

# --- Right Hand Section ---
st.header("📍 Right Hand Gait Data")
## refer hand as right hand (VV imp)

right_excel = st.file_uploader("Upload Right Hand Excel File (.xlsx)", type=["xlsx"], key="right_xlsx")
right_cell = "I2"

hand_input = ""
if right_excel and right_cell:
    extracted = extract_cell_value(right_excel, right_cell)
    if extracted:
        hand_input = extracted
        st.success("✅ Data extracted from cell.")
    else:
        st.error("❌ Could not extract from specified cell.")

hand_input = st.text_area("Comma-separated 300 values (Right Hand)", value=hand_input, height=150)
# hand_input = st.text_area("Enter 300 comma-separated values for right hand sensor:")

# Right Hand Image Upload
st.subheader("Upload Right Hand Plot")
right_hand_img = st.file_uploader("Choose an image of right hand gait analysis", type=["jpg", "jpeg", "png"], key="right_hand")
if right_hand_img is not None:
    image = Image.open(right_hand_img)
    st.image(image, caption="Uploaded Right Hand Image", width=300)
    
    # Convert image to bytes and store in session state
    right_hand_image_bytes = image_to_bytes(image)
    st.session_state["right_hand_image_bytes"] = right_hand_image_bytes


if st.button("Analyze Right Hand"):
    hand_array, error = parse_input(hand_input)
    if error:
        st.error(error)
    else:
        prediction = hand_model.predict(hand_array)[0]
        result = "🟢 Healthy" if prediction == 0 else "🔴 Parkinson"
        st.success(f"**Right Hand Gait Prediction:** {result}")

        # Save to session   
        st.session_state["gait_hand_result"] = result
        st.session_state["gait_hand_data"] = hand_array